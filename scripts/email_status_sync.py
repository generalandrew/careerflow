#!/usr/bin/env python3
"""
email_status_sync.py - Read recent emails for application status updates and
sync them to applications.xlsx + per-application metadata.json.

Uses IMAP. Requires user to configure their email credentials in
~/.careerflow_email_config.json (gitignored, never committed).

Config format:
    {
        "imap_server": "imap.gmail.com",
        "imap_port": 993,
        "email_address": "user@example.com",
        "password_or_app_password": "app-specific-password",
        "scan_folders": ["INBOX"],
        "scan_days_back": 14
    }

For Gmail, use an App Password (myaccount.google.com -> Security -> App passwords).
For Outlook 365, generate an app password at security.microsoft.com.

Match patterns (built in, customizable):
    - "Thank you for applying" -> confirm Status = Applied
    - "Unfortunately" or "we have decided" or "not moving forward" -> Status = Rejected
    - "Schedule" or "interview" or "phone screen" -> Status = Interviewing
    - "offer" or "we'd love to have you" -> Status = Offer

Usage:
    python3 scripts/email_status_sync.py [--dry-run] [--config <path>]
"""
import argparse
import datetime
import email
import imaplib
import json
import re
import sys
from email.header import decode_header
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
APP_INDEX = ROOT / "applications.xlsx"
DEFAULT_CONFIG = Path.home() / ".careerflow_email_config.json"


STATUS_PATTERNS = [
    (re.compile(r"\b(thank you for applying|application received|we have received your)\b", re.IGNORECASE),
     "Applied", "Application receipt confirmation"),
    (re.compile(r"\b(unfortunately|we have decided|not moving forward|not be moving|we will not|other candidates|after careful consideration)\b", re.IGNORECASE),
     "Rejected", "Rejection notice"),
    (re.compile(r"\b(schedule|phone screen|initial conversation|invite to interview|next steps)\b", re.IGNORECASE),
     "Interviewing", "Interview invitation or schedule request"),
    (re.compile(r"\b(offer letter|we'd love to have you|excited to extend|formally offer)\b", re.IGNORECASE),
     "Offer", "Offer extended"),
]


def decode_str(s):
    if isinstance(s, bytes):
        return s.decode("utf-8", errors="replace")
    return s


def get_email_body(msg):
    parts = []
    for part in msg.walk():
        ctype = part.get_content_type()
        if ctype == "text/plain":
            try:
                parts.append(part.get_payload(decode=True).decode("utf-8", errors="replace"))
            except Exception:
                pass
    return "\n".join(parts)


def match_company(subject, body, applications_index):
    """Try to associate an email with an application by matching company name in subject or body."""
    text = (subject + " " + body).lower()
    matches = []
    for app in applications_index:
        co = (app.get("Company") or "").strip()
        if not co:
            continue
        if co.lower() in text:
            matches.append(app)
    return matches


def load_applications():
    try:
        import openpyxl
    except ImportError:
        sys.stderr.write("openpyxl required\n")
        return []
    if not APP_INDEX.exists():
        return []
    wb = openpyxl.load_workbook(APP_INDEX)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    apps = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        apps.append(dict(zip(headers, row)))
    return apps


def update_application_status(folder, new_status, source_note):
    """Update metadata.json + applications.xlsx for one application."""
    try:
        import openpyxl
    except ImportError:
        return
    # metadata.json
    meta_path = ROOT / folder / "metadata.json"
    if meta_path.exists():
        meta = json.loads(meta_path.read_text())
        meta["status"] = new_status
        meta["last_touch"] = datetime.date.today().isoformat()
        meta.setdefault("status_history", []).append({
            "date": datetime.date.today().isoformat(),
            "status": new_status,
            "source": source_note,
        })
        meta_path.write_text(json.dumps(meta, indent=2))
    # xlsx
    wb = openpyxl.load_workbook(APP_INDEX)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2):
        if row[headers.index("Folder")].value == folder:
            row[headers.index("Status")].value = new_status
            row[headers.index("Last Touch")].value = datetime.date.today().isoformat()
            break
    wb.save(APP_INDEX)


def fetch_recent_emails(cfg, days_back):
    M = imaplib.IMAP4_SSL(cfg["imap_server"], cfg.get("imap_port", 993))
    M.login(cfg["email_address"], cfg["password_or_app_password"])
    since = (datetime.date.today() - datetime.timedelta(days=days_back)).strftime("%d-%b-%Y")
    messages = []
    for folder in cfg.get("scan_folders", ["INBOX"]):
        M.select(folder)
        typ, data = M.search(None, f'(SINCE {since})')
        for num in data[0].split():
            typ, msg_data = M.fetch(num, '(RFC822)')
            msg = email.message_from_bytes(msg_data[0][1])
            subject = decode_str(decode_header(msg["subject"])[0][0]) if msg["subject"] else ""
            body = get_email_body(msg)
            messages.append({"subject": subject, "body": body, "from": msg.get("From", "")})
    M.logout()
    return messages


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", default=str(DEFAULT_CONFIG))
    ap.add_argument("--dry-run", action="store_true", help="Show what would be updated, don't write")
    args = ap.parse_args()

    cfg_path = Path(args.config).expanduser()
    if not cfg_path.exists():
        sys.stderr.write(
            f"Config not found at {cfg_path}\n"
            f"Create it with the schema in this script's docstring.\n"
            f"Add {cfg_path.name} to your global gitignore.\n"
        )
        return 2
    cfg = json.loads(cfg_path.read_text())

    apps = load_applications()
    if not apps:
        sys.stderr.write("No applications found in applications.xlsx, nothing to sync.\n")
        return 0

    print(f"Scanning emails (last {cfg.get('scan_days_back', 14)} days)...")
    emails = fetch_recent_emails(cfg, cfg.get("scan_days_back", 14))
    print(f"  Fetched {len(emails)} messages.")

    updates = []
    for em in emails:
        subj = em["subject"]
        body = em["body"]
        full_text = subj + "\n" + body
        new_status = None
        source_note = ""
        for pattern, status, note in STATUS_PATTERNS:
            if pattern.search(full_text):
                new_status = status
                source_note = note
                break
        if not new_status:
            continue
        matched_apps = match_company(subj, body, apps)
        if not matched_apps:
            continue
        for app in matched_apps:
            updates.append({
                "company": app["Company"],
                "role": app.get("Role"),
                "folder": app["Folder"],
                "current_status": app.get("Status"),
                "new_status": new_status,
                "email_subject": subj,
                "source_note": source_note,
            })

    if not updates:
        print("No status updates detected.")
        return 0

    print(f"\n{len(updates)} status updates detected:")
    for u in updates:
        print(f"  {u['company']} ({u['role']}): {u['current_status']} -> {u['new_status']}  [{u['source_note']}]")
        print(f"    Email: \"{u['email_subject'][:80]}\"")

    if args.dry_run:
        print("\n--dry-run, no changes written.")
        return 0

    print("\nApplying updates...")
    for u in updates:
        update_application_status(u["folder"], u["new_status"], u["source_note"])
    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
