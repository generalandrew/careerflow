#!/usr/bin/env python3
"""
write_summary.py - Write posting_summary.md and update applications.xlsx
                   with per-application catalog data.

Inputs are passed as a single JSON document on stdin:

  {
    "app_dir": "/abs/path/to/application/folder",
    "company": "Acme",
    "role": "Senior Solutions Engineer",
    "url": "https://...",
    "company_summary": "Acme is a leader in...",
    "responsibilities_concise": ["...", "..."],
    "qualifications": ["...", "..."],         (optional)
    "match_snapshot": ["...", "..."],         (optional)
    "talking_points": ["...", "..."],         (optional)
    "status": "Draft" | "Applied"             (optional, default Draft)
  }

Updates posting.json with company_summary + responsibilities_concise.
Writes posting_summary.md.
Appends/updates row in applications.xlsx (matched by Folder).
"""
import sys, os, json, datetime
from pathlib import Path


def main():
    payload = json.load(sys.stdin)
    app_dir = Path(payload["app_dir"]).resolve()
    if not app_dir.exists():
        sys.stderr.write(f"app_dir not found: {app_dir}\n"); return 2

    # Update posting.json
    posting_path = app_dir / "posting.json"
    posting = json.loads(posting_path.read_text()) if posting_path.exists() else {}
    posting["company_summary"] = payload.get("company_summary", "")
    posting["responsibilities_concise"] = payload.get("responsibilities_concise", [])
    if payload.get("company") and not posting.get("company"):
        posting["company"] = payload["company"]
    if payload.get("role") and not posting.get("title"):
        posting["title"] = payload["role"]
    if payload.get("url") and not posting.get("url"):
        posting["url"] = payload["url"]
    posting_path.write_text(json.dumps(posting, indent=2))

    # Write posting_summary.md
    comp = posting.get("compensation", {})
    comp_line = ""
    if comp.get("salary_min"):
        rng = f"${comp['salary_min']:,} - ${comp['salary_max']:,}" if comp.get("salary_max") else f"${comp['salary_min']:,}+"
        extras = []
        if comp.get("bonus_or_commission"): extras.append("bonus/commission")
        if comp.get("equity"): extras.append("equity")
        comp_line = rng + (" + " + ", ".join(extras) if extras else "")

    md = []
    md.append(f"# {payload.get('company','?')} - {payload.get('role','?')}\n")
    if payload.get("url"):
        md.append(f"**Source:** {payload['url']}  ")
    md.append(f"**Captured:** {datetime.date.today().isoformat()}  ")
    if posting.get("location"):
        md.append(f"**Location:** {posting['location']}  ")
    if comp_line:
        md.append(f"**Compensation:** {comp_line}  ")
    if comp.get("benefits"):
        md.append(f"**Benefits:** {', '.join(comp['benefits'])}  ")
    md.append("")
    md.append("## Company Summary")
    md.append(payload.get("company_summary","") + "\n")
    md.append("## Concise Responsibilities")
    for b in payload.get("responsibilities_concise", []):
        md.append(f"- {b}")
    md.append("")
    if payload.get("qualifications"):
        md.append("## Required / Preferred Qualifications")
        for b in payload["qualifications"]:
            md.append(f"- {b}")
        md.append("")
    if payload.get("match_snapshot"):
        md.append("## Match Snapshot")
        for b in payload["match_snapshot"]:
            md.append(f"- {b}")
        md.append("")
    if payload.get("talking_points"):
        md.append("## Talking Points")
        for b in payload["talking_points"]:
            md.append(f"- {b}")
        md.append("")
    (app_dir / "posting_summary.md").write_text("\n".join(md))

    # Append/update applications.xlsx
    try:
        import openpyxl
    except ImportError:
        sys.stderr.write("openpyxl required\n"); return 0

    root = app_dir.parent.parent  # applications/<folder>/.. -> root
    xlsx = root / "applications.xlsx"
    if not xlsx.exists():
        sys.stderr.write(f"applications.xlsx not found at {xlsx}, skipping index update\n")
        return 0
    wb = openpyxl.load_workbook(xlsx)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    folder_rel = str(app_dir.relative_to(root))
    matched = None
    for row in ws.iter_rows(min_row=2):
        if row[headers.index("Folder")].value == folder_rel:
            matched = row
            break

    today = datetime.date.today().isoformat()
    status = payload.get("status", "Draft")
    row_data = {
        "Date Applied": today,
        "Company": payload.get("company", ""),
        "Role": payload.get("role", ""),
        "URL": payload.get("url", ""),
        "Status": status,
        "Compensation": comp_line,
        "Folder": folder_rel,
        "Contact Name": "",
        "Contact Email": "",
        "Last Touch": today,
        "Notes": "",
    }
    if matched:
        for h, v in row_data.items():
            if h in headers and v:
                matched[headers.index(h)].value = v
    else:
        ws.append([row_data.get(h, "") for h in headers])
    wb.save(xlsx)
    sys.stdout.write(f"Updated {app_dir/'posting_summary.md'} and applications.xlsx\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
