#!/usr/bin/env python3
"""
pipeline_digest.py - Generate a pipeline digest report covering:
  - Applications submitted this week
  - Status changes
  - Applications needing followup (>N days since Last Touch)
  - Interview invitations awaiting response
  - Recent hot leads from discovery scans

Output: a Markdown digest file in the workspace root.

Usage:
    python3 scripts/pipeline_digest.py [--days 7] [--out digest_<date>.md]
"""
import argparse
import datetime
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
APP_INDEX = ROOT / "applications.xlsx"


def load_applications():
    try:
        import openpyxl
    except ImportError:
        sys.stderr.write("openpyxl required\n")
        return []
    if not APP_INDEX.exists():
        return []
    wb = openpyxl.load_workbook(APP_INDEX)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]


def days_ago(date_str):
    if not date_str:
        return None
    try:
        d = datetime.date.fromisoformat(str(date_str))
        return (datetime.date.today() - d).days
    except Exception:
        return None


def find_recent_candidate_files(limit=3):
    files = sorted(ROOT.glob("candidates_v*.md"), key=lambda p: p.stat().st_mtime, reverse=True)
    return files[:limit]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--days", type=int, default=7, help="Recent window in days for new applications and status changes")
    ap.add_argument("--followup-threshold", type=int, default=7, help="Days since Last Touch to flag for followup")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    today = datetime.date.today()
    apps = load_applications()
    if not apps:
        print("No applications found.")
        return 0

    # Bucket
    submitted_recent = []
    status_changes_recent = []  # rows where Last Touch is recent and Status != Applied (implies a change)
    needs_followup = []
    interviewing = []
    offers = []
    for a in apps:
        last_touch_days = days_ago(a.get("Last Touch"))
        date_applied_days = days_ago(a.get("Date Applied"))
        status = (a.get("Status") or "").strip()
        if date_applied_days is not None and date_applied_days <= args.days:
            submitted_recent.append(a)
        if last_touch_days is not None and last_touch_days <= args.days and status not in ("Draft", "Applied"):
            status_changes_recent.append(a)
        if status == "Applied" and last_touch_days is not None and last_touch_days >= args.followup_threshold:
            needs_followup.append(a)
        if status == "Interviewing":
            interviewing.append(a)
        if status == "Offer":
            offers.append(a)

    # Recent candidate files
    candidate_files = find_recent_candidate_files(3)

    # Build digest
    md = []
    md.append(f"# Pipeline Digest, {today.isoformat()}")
    md.append("")
    md.append(f"_Window, last {args.days} days. Followup threshold, {args.followup_threshold} days since last touch._\n")

    md.append(f"## Summary")
    md.append(f"- Submitted in last {args.days} days, {len(submitted_recent)}")
    md.append(f"- Status changes in last {args.days} days, {len(status_changes_recent)}")
    md.append(f"- Applications needing followup, {len(needs_followup)}")
    md.append(f"- Currently interviewing, {len(interviewing)}")
    md.append(f"- Active offers, {len(offers)}")
    md.append("")

    if offers:
        md.append("## Active Offers (highest priority)")
        for a in offers:
            md.append(f"- {a['Company']} - {a.get('Role','?')} (folder: {a['Folder']})")
        md.append("")

    if interviewing:
        md.append("## Currently Interviewing")
        for a in interviewing:
            md.append(f"- {a['Company']} - {a.get('Role','?')} (last touch: {a.get('Last Touch','?')})")
        md.append("")

    if needs_followup:
        md.append("## Applications Needing Followup")
        md.append("Sorted by age (oldest first). Run `run followup scan` to generate draft emails.")
        sorted_followup = sorted(needs_followup, key=lambda a: a.get("Last Touch") or "")
        for a in sorted_followup:
            d = days_ago(a.get("Last Touch")) or 0
            md.append(f"- {a['Company']} - {a.get('Role','?')}, {d} days since last touch")
        md.append("")

    if status_changes_recent:
        md.append(f"## Status Changes (last {args.days} days)")
        for a in status_changes_recent:
            md.append(f"- {a['Company']} - {a.get('Role','?')}, now {a.get('Status','?')}")
        md.append("")

    if submitted_recent:
        md.append(f"## New Submissions (last {args.days} days)")
        for a in submitted_recent:
            md.append(f"- {a['Company']} - {a.get('Role','?')} (status: {a.get('Status','?')})")
        md.append("")

    if candidate_files:
        md.append("## Recent Discovery Scans (top 3 candidate files)")
        for f in candidate_files:
            mtime = datetime.date.fromtimestamp(f.stat().st_mtime).isoformat()
            md.append(f"- [{f.name}]({f.name}) (generated {mtime})")
        md.append("")

    md.append("---")
    md.append("Recommended actions:")
    actions = []
    if offers:
        actions.append("Review offers and respond by deadline.")
    if interviewing:
        actions.append("Prepare for upcoming interviews. Use `build interview prep for [Company]`.")
    if needs_followup:
        actions.append(f"Run `run followup scan` to generate draft followup emails for {len(needs_followup)} stale applications.")
    if not submitted_recent and not offers and not interviewing:
        actions.append("Pipeline is quiet. Consider running a discovery scan to surface fresh candidates.")
    for a in actions:
        md.append(f"- {a}")

    out_path = ROOT / (args.out or f"digest_{today.isoformat()}.md")
    out_path.write_text("\n".join(md))
    print(f"Wrote {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
