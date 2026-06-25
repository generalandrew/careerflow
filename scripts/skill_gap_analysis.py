#!/usr/bin/env python3
"""
skill_gap_analysis.py - Analyze the user's application outcomes to surface
patterns: which roles/skills get rejected most, which get callbacks most,
where the user's targeting may need to shift.

Reads applications.xlsx + per-application posting.json + tailored.json
across the user's history. Computes:

  - Rejection-to-application ratio per target_job_type
  - Rejection-to-application ratio per industry
  - Skills/keywords present in posting.json of rejected applications but
    absent from the user's experience.json (potential skill gap)
  - Skills present in callback/offer postings but not yet emphasized in
    the user's tailored resumes
  - Recommendation summary

Usage:
    python3 scripts/skill_gap_analysis.py [--out skill_gap_<date>.md]
"""
import argparse
import collections
import datetime
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
APP_INDEX = ROOT / "applications.xlsx"
EXP_PATH = ROOT / "master" / "experience.json"


def load_applications():
    try:
        import openpyxl
    except ImportError:
        sys.stderr.write("openpyxl required\n")
        return []
    if not APP_INDEX.exists():
        return []
    wb = openpyxl.load_workbook(APP_INDEX)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]


def load_posting(folder):
    p = ROOT / folder / "posting.json"
    if p.exists():
        try:
            return json.loads(p.read_text())
        except Exception:
            return {}
    return {}


def load_experience():
    if EXP_PATH.exists():
        return json.loads(EXP_PATH.read_text())
    return {}


def classify_role_family(role_title, targeting):
    """Best-effort matching of a posting role to one of the user's target_job_types."""
    if not role_title:
        return None
    rt = role_title.lower()
    for fam in targeting.get("target_job_types", []):
        if fam.lower() in rt:
            return fam
    return "Other"


def all_user_skills(exp):
    """Flatten all skills + role bullets into a single text blob for keyword matching."""
    chunks = []
    skills = exp.get("skills", {})
    if isinstance(skills, dict):
        for cat, items in skills.items():
            chunks.append(cat)
            chunks.extend(items if isinstance(items, list) else [])
    elif isinstance(skills, list):
        chunks.extend(skills)
    for role in exp.get("roles", []):
        chunks.append(role.get("title", ""))
        for b in role.get("bullets", []):
            chunks.append(b if isinstance(b, str) else b.get("text", ""))
    return " ".join(chunks).lower()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    apps = load_applications()
    exp = load_experience()
    if not apps:
        print("No applications to analyze.")
        return 0
    if not exp:
        sys.stderr.write("experience.json not found, skill gap analysis requires it\n")
        return 2

    targeting = exp.get("targeting", {})
    skill_blob = all_user_skills(exp)

    # Bucket apps
    by_status = collections.defaultdict(list)
    for a in apps:
        by_status[(a.get("Status") or "Unknown").strip()].append(a)

    total = len(apps)
    rejected = by_status.get("Rejected", [])
    offered = by_status.get("Offer", []) + by_status.get("Accepted", [])
    interviewing = by_status.get("Interviewing", [])

    # Per-family stats
    fam_stats = collections.defaultdict(lambda: {"applied": 0, "rejected": 0, "interviewing": 0, "offered": 0})
    for a in apps:
        fam = classify_role_family(a.get("Role", ""), targeting)
        fam_stats[fam]["applied"] += 1
        status = (a.get("Status") or "").strip()
        if status == "Rejected":
            fam_stats[fam]["rejected"] += 1
        if status == "Interviewing":
            fam_stats[fam]["interviewing"] += 1
        if status in ("Offer", "Accepted"):
            fam_stats[fam]["offered"] += 1

    # Keyword gap: keywords in rejected postings not in user's skill blob
    rejected_keywords = collections.Counter()
    for a in rejected:
        posting = load_posting(a.get("Folder", ""))
        for kw in posting.get("keywords", []):
            if kw.lower() not in skill_blob:
                rejected_keywords[kw] += 1

    # Keyword strengths: keywords in offered/interviewing postings present in skill blob
    strong_keywords = collections.Counter()
    for a in offered + interviewing:
        posting = load_posting(a.get("Folder", ""))
        for kw in posting.get("keywords", []):
            if kw.lower() in skill_blob:
                strong_keywords[kw] += 1

    # Build report
    md = []
    today = datetime.date.today()
    md.append(f"# Skill Gap Analysis, {today.isoformat()}")
    md.append("")
    md.append(f"_Total applications analyzed: {total}_")
    md.append(f"_Rejected: {len(rejected)}, Interviewing: {len(interviewing)}, Offered/Accepted: {len(offered)}_")
    md.append("")

    md.append("## Per Job-Type Family Stats")
    md.append("| Family | Applied | Rejected | Interviewing | Offered |")
    md.append("|--------|---------|----------|--------------|---------|")
    for fam, s in sorted(fam_stats.items(), key=lambda kv: -kv[1]["applied"]):
        md.append(f"| {fam} | {s['applied']} | {s['rejected']} | {s['interviewing']} | {s['offered']} |")
    md.append("")

    md.append("## Potential Skill Gaps (keywords in rejected postings, absent from your profile)")
    if rejected_keywords:
        md.append("Top 20 keywords appearing most often in rejected postings that do not show up in your skills or roles:")
        for kw, count in rejected_keywords.most_common(20):
            md.append(f"- **{kw}** ({count} occurrences)")
    else:
        md.append("No gaps detected, or insufficient rejected applications to analyze.")
    md.append("")

    md.append("## Strength Signal (keywords in callback/offer postings present in your profile)")
    if strong_keywords:
        md.append("Top 20 keywords your profile already covers and that correlate with callbacks:")
        for kw, count in strong_keywords.most_common(20):
            md.append(f"- **{kw}** ({count} occurrences)")
    else:
        md.append("Not enough callback signal yet to identify strengths.")
    md.append("")

    md.append("## Recommendations")
    recs = []
    # Family with worst rejection rate
    if fam_stats:
        worst = sorted(fam_stats.items(), key=lambda kv: -(kv[1]["rejected"] / max(kv[1]["applied"], 1)))[0]
        if worst[1]["applied"] >= 5 and worst[1]["rejected"] / worst[1]["applied"] > 0.6:
            recs.append(
                f"**{worst[0]}** has a high rejection rate "
                f"({worst[1]['rejected']}/{worst[1]['applied']}). "
                "Consider whether this target family is the right fit, "
                "or whether the resume positioning needs revision."
            )
    # Best converting family
    if fam_stats:
        best = sorted(
            fam_stats.items(),
            key=lambda kv: -(kv[1]["interviewing"] + kv[1]["offered"]) / max(kv[1]["applied"], 1)
        )[0]
        if best[1]["applied"] >= 5 and (best[1]["interviewing"] + best[1]["offered"]) / best[1]["applied"] > 0.3:
            recs.append(
                f"**{best[0]}** has the strongest callback rate. "
                "Consider doubling down on this family in your next discovery scans."
            )
    if rejected_keywords:
        top_gaps = [k for k, _ in rejected_keywords.most_common(5)]
        recs.append(
            f"Top recurring gaps: {', '.join(top_gaps)}. "
            "Consider building hands-on experience, a certification, or a project that demonstrates these."
        )
    if not recs:
        recs.append("Insufficient data for actionable recommendations. Re-run after 20+ applications.")
    for r in recs:
        md.append(f"- {r}")
    md.append("")

    out_path = ROOT / (args.out or f"skill_gap_{today.isoformat()}.md")
    out_path.write_text("\n".join(md))
    print(f"Wrote {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
